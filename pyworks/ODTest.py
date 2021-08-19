# Get the Current Working Folder
import os

os.getcwd()

# Importing openpyxl library
import openpyxl

# Loading the Workbook and printing the object
wb = openpyxl.load_workbook('PROD-091.xlsx')
print('Openpyxl Class: ', type(wb))

# Sheets in Excel Workbook
print('Sheets in the Excel file: ', wb.sheetnames)

# Active Sheet in the Workbook
print('Active Worksheet: ', wb.active)

actvSheet = wb.active
sheet = wb.get_sheet_by_name('Sheet1')

ODData = {}

for row in range(2, sheet.max_row + 1):
    headernetwork = sheet['A' + str(row).value]
    ipv4address = sheet['K' + str(row).value]
    ipv4netmask = sheet['L' + str(row).value]
    updateversion = sheet['B' + str(row).value]
    deployment = sheet['C' + str(row).value]
    datacenter = sheet['D' + str(row).value]
    tenant = sheet['E' + str(row).value]
    telco12 = sheet['F' + str(row).value]
    networktype = sheet['G' + str(row).value]
    domaintype = sheet['H' + str(row).value]
    domainname = sheet['I' + str(row).value]
    zone = sheet['J' + str(row).value]
    subnetname = sheet['K' + str(row).value]
    dhcp = sheet['L' + str(row).value]
    ssname = sheet['M' + str(row).value]
    vnfid = sheet['N' + str(row).value]
    pnfvnf = sheet['O' + str(row).value]
    bgpv4 = sheet['P' + str(row).value]
    sriovovs = sheet['A' + str(row).value]
    comment = sheet['A' + str(row).value]
    physnetid = sheet['A' + str(row).value]
    sriovtrunk = sheet['A' + str(row).value]
    l2duplex = sheet['A' + str(row).value]
    vlanid = sheet['A' + str(row).value]
    ipv4apstart = sheet['A' + str(row).value]
    ipv4apend = sheet['A' + str(row).value]
    ipv6apstart = sheet['A' + str(row).value]
    # ipv6apend = sheet['A' + str(row).value]
    # physnetid2 = sheet['A' + str(row).value]
    # ipv4apstart2 = sheet['A' + str(row).value]
    # ipv4apend2 = sheet['A' + str(row).value]
    # ipv6apstart2 = sheet['A' + str(row).value]
    # ipv6apend2 = sheet['A' + str(row).value]

fobj = open('IPV4IPAMImport.csv', 'w')

#
# header-network - A2; address* - K2; netmask* - L2; EA-Update_Version - B2; EA-Deployment - C2; EA-DataCenter - D2
# EA-Tenant - E2; EA-Telco 1_2 - F2; EA-Network Type - G2; EA-Domain Type - G2; EA-Domain Name - H2; EA-ZONE - I2  
# EA-Subnet name - J2; EA-DHCP - O2; EA-NAME - P2; EA-VNFID - Q2; EA-PNF/VNF - R2; EA-BGPv4 - S2; EA-SRIOV/OVS - T2 
# Comment - U2; EA-Physnet ID - ""; EA-SR-IOV Trunk - ""; EA-L2 Duples - ""; EA-VLAN ID - ""; EA-IPv4 AP Start - ""
# EA-IPv4 AP End - ""; EA-IPv6 AP Start - ""; EA-IPv6 AP End - ""; EA-Physnet ID2 - ""; EA-IPv4 AP Start2 - "" 
# EA-IPv4 AP End2 - ""; EA-IPv6 AP Start2 - ""; EA-IPv6 AP End2 - ""
#

fobj.write(
    'header-network, address*, netmask*, EA-Update_Version, EA-Deployment, EA-DataCenter, EA-Tenant, EA-Telco 1_2, '
    'EA-Network Type, EA-Domain Type, EA-Domain Name, EA-ZONE, EA-Subnet name, EA-DHCP, EA-NAME, EA-VNFID, EA-PNF/VNF, '
    'EA-BGPv4, EA-SRIOV/OVS, Comment, EA-Physnet ID, EA-SR-IOV Trunk, EA-L2 Duplex, EA-VLAN ID, EA-IPv4 AP Start, '
    'EA-IPv4 AP End, EA-IPv6 AP Start, EA-IPv6 AP End, EA-Physnet ID2, EA-IPv4 AP Start2, EA-IPv4 AP End2, '
    'EA-IPv6 AP Start2, EA-IPv6 AP End2 \n')

old = [
    (actvSheet['A2'].value, actvSheet['K2'].value, actvSheet['L2'].value, actvSheet['B2'].value, actvSheet['C2'].value,
     actvSheet['D2'].value, actvSheet['E2'].value, actvSheet['F2'].value, (actvSheet['G2'].value, actvSheet['G2'].value,
                                                                           actvSheet['H2'].value, actvSheet['I2'].value,
                                                                           actvSheet['J2'].value, actvSheet['O2'].value,
                                                                           actvSheet['P2'].value,
                                                                           actvSheet['Q2'].value),
     actvSheet['R2'].value, actvSheet['S2'].value, actvSheet['T2'].value, actvSheet['U2'].value,
     'test', '', '', '', '', '', '', '', '', '', '', '', '')]

for od in old:
    headernetwork = od[0]
    ipv4address = od[1]
    ipv4netmask = od[2]
    updateversion = od[3]
    deployment = od[4]
    datacenter = od[5]
    tenant = od[6]
    telco12 = od[7]
    networktype = od[8]
    domaintype = od[9]
    domainname = od[10]
    zone = od[11]
    subnetname = od[12]
    dhcp = od[13]
    ssname = od[14]
    vnfid = od[15]
    pnfvnf = od[16]
    bgpv4 = od[17]
    sriovovs = od[18]
    comment = od[19]
    physnetid = od[20]
    sriovtrunk = od[21]
    l2duplex = od[22]
    vlanid = od[23]
    ipv4apstart = od[24]
    ipv4apend = od[25]
    ipv6apstart = od[26]
    # ipv6apend = od[27]
    # physnetid2 = od[28]
    # ipv4apstart2 = od[29]
    # ipv4apend2 = od[30]
    # ipv6apstart2 = od[31]
    # ipv6apend2 = od[32]

    # fobj.write(str(headernetwork) + "," + str(ipv4address) + "," + str(ipv4netmask) + "," + str(updateversion) + "," +
    #            str(deployment) + "," + str(datacenter) + "," + str(tenant) + "," + str(telco12) + "," + str(networktype) + "," +
    #            str(domaintype) + "," + str(zone) + "," + str(subnetname) + "," + str(dhcp) + "," + str(ssname) + "," +
    #            str(vnfid) + "," + str(pnfvnf) + "," + str(bgpv4) + "," + str(sriovovs) + "," + str(comment) + "," +
    #            str(physnetid) + "," + str(sriovtrunk) + "," + str(l2duplex) + "," + str(vlanid) + "," + str(ipv4apstart) + "," +
    #            str(ipv4apend) + "," + str(ipv6apstart) + "," + str(ipv6apend) + "," + str(physnetid2) + "," + str(ipv4apstart2) + "," +
    #            str(ipv4apend2) + "," + str(ipv6apstart2) + "," + str(ipv6apend2))

    fobj.write(str(headernetwork) + "," + str(ipv4address) + "," + str(ipv4netmask) + "," + str(updateversion) + "," +
               str(deployment) + "," + str(datacenter) + "," + str(tenant) + "," + str(telco12) + "," + str(
        networktype) + "," +
               str(domaintype) + "," + str(zone) + "," + str(subnetname) + "," + str(dhcp) + "," + str(ssname) + "," +
               str(vnfid) + "," + str(pnfvnf) + "," + str(bgpv4) + "," + str(sriovovs) + "," + str(comment) + "," +
               str(physnetid) + "," + str(sriovtrunk) + "," + str(l2duplex) + "," + str(vlanid) + "," + str(
        ipv4apstart) + "," +
               str(ipv4apend), str(ipv6apstart))

# Printing the Column Headers & Values
print("Printing CSV Headers")
print('header-network, address*, netmask, EA-Update_Version, EA-Deployment, EA-DataCenter, EA-Tenant, EA-Telco 1_2, '
      'EA-Network Type, EA-Domain Type, EA-Domain Name, EA-ZONE, EA-Subnet name, EA-DHCP, EA-NAME, EA-VNFID, EA-PNF/VNF, '
      'EA-BGPv4, EA-SRIOV/OVS, Comment, EA-Physnet ID, EA-SR-IOV Trunk, EA-L2 Duplex, EA-VLAN ID, EA-IPv4 AP Start, '
      'EA-IPv4 AP End, EA-IPv6 AP Start, EA-IPv6 AP End, EA-Physnet ID2, EA-IPv4 AP Start2, EA-IPv4 AP End2, '
      'EA-IPv6 AP Start2, EA-IPv6 AP End2 \n')

print("End of Printing")
