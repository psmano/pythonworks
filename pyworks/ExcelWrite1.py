from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([11, 22, 33])

# Python types will automatically be converted
import datetime
ws['A3'] = datetime.datetime.now()
ws['A4'] = datetime.datetime.date(datetime.datetime.now())
ws['A5'] = datetime.datetime.time(datetime.datetime.now())
ws['A6'] = datetime.datetime.today()
ws['A7'] = "This is a test"
ws['A8'] = 100
ws['A9'] = 200
ws['A10'] = "=SUM(A8,A9)"


# Save the file
wb.save("sample.xlsx")

