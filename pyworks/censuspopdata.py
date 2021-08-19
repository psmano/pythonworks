import openpyxl, pprint
import os

print('Opening Workbook')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')
countyData = {}

print('Reading rows...')

for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

countyData.setdefault(state, {})
countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
countyData[state][county]['tracts'] += 1
countyData[state][county]['pop'] += int(pop)

print("Writing Results...")
resultFile = open('census2010.py', 'w')
for row in range(2, sheet.max_row + 1):

resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')

#os.chdir('C:\\Manohar\\Python')
import census2010


#census2010.allData['AK']['Anchorage']

anchoragePop = census2010.allData['WY']['Weston']['pop']

print('The 2010 population of Weston was ' + str(anchoragePop))



